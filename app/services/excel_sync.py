"""Excel synchronization — parse-only + apply stages.

Core rules (see project spec):
  1. Match by `excel_row_id` when present, else `(brand, model, year)`.
  2. Parse stage NEVER writes to DB; it returns a diff and caches it under a token.
  3. Apply stage runs in a single transaction:
       - new rows   → insert with status=pending, source=excel → notify admins
       - updated    → update + audit log (price changes logged separately)
       - removed    → status=archived (NEVER hard delete) → warn admins
  4. Malformed file → reject at parse stage with warnings.
"""
from __future__ import annotations

import json
import uuid
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.cache import cache_get, cache_set, get_redis, invalidate_cars_cache
from app.models import (
    AuditAction,
    AuditLog,
    Car,
    CarSource,
    CarStatus,
    ExcelImportLog,
)

REQUIRED_COLUMNS = {"brand", "model", "year", "price"}
OPTIONAL_COLUMNS = {
    "mileage",
    "color",
    "gearbox",
    "fuel_type",
    "location",
    "excel_row_id",
    "body_status",
    "description",
}
ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

TOKEN_TTL_SECONDS = 60 * 60  # 1 hour


# ───────────────────────── Parse ─────────────────────────


def _normalize_header(cell: Any) -> str:
    return str(cell).strip().lower().replace(" ", "_") if cell is not None else ""


def _coerce_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _coerce_price(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _row_key(row: dict[str, Any]) -> str:
    if row.get("excel_row_id"):
        return f"eid:{row['excel_row_id']}"
    return f"bmy:{str(row.get('brand', '')).strip().lower()}|{str(row.get('model', '')).strip().lower()}|{row.get('year')}"


def parse_workbook(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Return (rows, warnings). Raises ValueError for fatal file-level problems."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open workbook: {exc}") from exc

    try:
        ws = wb.active
        if ws is None or ws.max_row < 1:
            raise ValueError("Workbook has no sheets or is empty")

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("Header row missing") from None

        headers = [_normalize_header(c) for c in header_row]
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError(f"Missing required columns: {sorted(missing)}")

        rows: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []
        seen_keys: dict[str, int] = {}

        for idx, raw in enumerate(rows_iter, start=2):
            if raw is None or all(c is None or c == "" for c in raw):
                continue
            record = {h: v for h, v in zip(headers, raw) if h in ALL_COLUMNS}

            # Field-level validation
            row_warnings: list[str] = []
            for col in REQUIRED_COLUMNS:
                if record.get(col) in (None, ""):
                    row_warnings.append(f"missing {col}")

            year = _coerce_int(record.get("year"))
            if year is None and record.get("year") not in (None, ""):
                row_warnings.append("invalid year")
            record["year"] = year

            price = _coerce_price(record.get("price"))
            if price is None and record.get("price") not in (None, ""):
                row_warnings.append("invalid price")
            elif price is not None and price <= 0:
                row_warnings.append("price must be > 0")
            record["price"] = price

            mileage = _coerce_int(record.get("mileage"))
            record["mileage"] = mileage

            if record.get("brand"):
                record["brand"] = str(record["brand"]).strip()
            if record.get("model"):
                record["model"] = str(record["model"]).strip()
            if record.get("excel_row_id") is not None:
                record["excel_row_id"] = str(record["excel_row_id"]).strip() or None

            if row_warnings:
                for issue in row_warnings:
                    warnings.append({"row": idx, "issue": issue})
                # Rows with required-field issues do NOT enter the diff
                continue

            key = _row_key(record)
            if key in seen_keys:
                warnings.append(
                    {"row": idx, "issue": f"duplicate key (seen on row {seen_keys[key]})"}
                )
                continue
            seen_keys[key] = idx
            record["_row_number"] = idx
            record["_key"] = key
            rows.append(record)

        return rows, warnings
    finally:
        wb.close()


# ───────────────────────── Diff ─────────────────────────


async def _load_existing_excel_cars(db: AsyncSession) -> list[Car]:
    stmt = select(Car).where(
        and_(
            Car.deleted_at.is_(None),
            Car.source == CarSource.excel,
            Car.status.in_([CarStatus.pending, CarStatus.published]),
        )
    )
    return list((await db.execute(stmt)).scalars().all())


def _db_key(car: Car) -> str:
    if car.excel_row_id:
        return f"eid:{car.excel_row_id}"
    return f"bmy:{car.brand.strip().lower()}|{car.model.strip().lower()}|{car.year}"


def _diff_row(row: dict[str, Any], car: Car) -> dict[str, Any]:
    changes: dict[str, Any] = {}
    for field in ("price", "mileage", "color", "gearbox", "fuel_type", "location", "body_status", "description"):
        if field not in row:
            continue
        incoming = row[field]
        existing = getattr(car, field, None)
        if field == "price" and incoming is not None and existing is not None:
            if Decimal(str(existing)) != Decimal(str(incoming)):
                changes[field] = {"old": str(existing), "new": str(incoming)}
        elif field in ("gearbox", "fuel_type") and existing is not None:
            if str(existing.value if hasattr(existing, "value") else existing) != str(incoming or ""):
                if incoming:
                    changes[field] = {"old": str(existing), "new": str(incoming)}
        elif (incoming or "") != (existing or ""):
            if incoming is not None:
                changes[field] = {"old": existing, "new": incoming}
    return changes


async def build_diff(
    db: AsyncSession, rows: list[dict[str, Any]], warnings: list[dict[str, Any]]
) -> dict[str, Any]:
    existing = await _load_existing_excel_cars(db)
    by_key = {_db_key(c): c for c in existing}

    new_cars: list[dict[str, Any]] = []
    updated_cars: list[dict[str, Any]] = []
    unchanged = 0
    matched_keys: set[str] = set()

    for row in rows:
        key = row["_key"]
        car = by_key.get(key)
        if not car:
            new_cars.append(
                {
                    "row_number": row["_row_number"],
                    "brand": row["brand"],
                    "model": row["model"],
                    "year": row["year"],
                    "price": str(row["price"]),
                    "excel_row_id": row.get("excel_row_id"),
                    "data": {k: (str(v) if isinstance(v, Decimal) else v) for k, v in row.items() if not k.startswith("_")},
                }
            )
            continue
        matched_keys.add(key)
        changes = _diff_row(row, car)
        if not changes:
            unchanged += 1
            continue
        updated_cars.append(
            {
                "row_number": row["_row_number"],
                "car_id": car.id,
                "brand": car.brand,
                "model": car.model,
                "year": car.year,
                "changes": changes,
            }
        )

    removed_cars = [
        {
            "car_id": c.id,
            "brand": c.brand,
            "model": c.model,
            "year": c.year,
            "price": str(c.price),
        }
        for key, c in by_key.items()
        if key not in matched_keys
    ]

    return {
        "new_cars": new_cars,
        "updated_cars": updated_cars,
        "removed_cars": removed_cars,
        "unchanged": unchanged,
        "warnings": warnings,
    }


# ───────────────────── Tokens (Redis) ─────────────────────


async def _stash_diff(
    file_path: Path, diff: dict[str, Any], rows: list[dict[str, Any]]
) -> str:
    token = uuid.uuid4().hex
    r = await get_redis()
    payload = {
        "file_path": str(file_path),
        "diff": diff,
        "rows": [
            {k: (str(v) if isinstance(v, Decimal) else v) for k, v in row.items()}
            for row in rows
        ],
    }
    await r.set(f"excel:diff:{token}", json.dumps(payload, default=str), ex=TOKEN_TTL_SECONDS)
    return token


async def load_diff(token: str) -> dict[str, Any] | None:
    r = await get_redis()
    raw = await r.get(f"excel:diff:{token}")
    if not raw:
        return None
    return json.loads(raw)


async def clear_diff(token: str) -> None:
    r = await get_redis()
    await r.delete(f"excel:diff:{token}")


# ───────────────────── Public entry points ─────────────────────


async def parse_and_stage(
    db: AsyncSession, file_path: Path, *, scheduled: bool = False
) -> tuple[dict[str, Any], str]:
    rows, warnings = parse_workbook(file_path)
    diff = await build_diff(db, rows, warnings)
    token = await _stash_diff(file_path, diff, rows)
    return diff, token


async def apply_diff(db: AsyncSession, token: str, user_id: int | None) -> dict[str, Any]:
    from datetime import datetime, timezone

    payload = await load_diff(token)
    if not payload:
        raise ValueError("Unknown or expired diff token")

    rows = payload["rows"]
    file_path = payload["file_path"]
    rows_by_key = {r["_key"]: r for r in rows}

    existing = await _load_existing_excel_cars(db)
    by_key = {_db_key(c): c for c in existing}

    added = 0
    updated = 0
    removed = 0

    # 1. New and updated
    for key, row in rows_by_key.items():
        incoming = {
            k: v
            for k, v in row.items()
            if not k.startswith("_") and k in ALL_COLUMNS
        }
        price = Decimal(str(incoming["price"])) if incoming.get("price") is not None else None
        year = int(incoming["year"]) if incoming.get("year") is not None else None
        mileage = int(incoming["mileage"]) if incoming.get("mileage") is not None else None

        if key in by_key:
            car = by_key[key]
            changed_fields: dict[str, Any] = {}
            old_price = car.price
            price_changed = False

            field_map: dict[str, Any] = {
                "price": price,
                "mileage": mileage,
                "color": incoming.get("color"),
                "location": incoming.get("location"),
                "body_status": incoming.get("body_status"),
                "description": incoming.get("description"),
            }
            for field, new_val in field_map.items():
                if new_val is None:
                    continue
                current = getattr(car, field)
                if field == "price" and current is not None and Decimal(str(current)) == new_val:
                    continue
                if field != "price" and (current or "") == (new_val or ""):
                    continue
                changed_fields[field] = {"old": current, "new": new_val}
                setattr(car, field, new_val)
                if field == "price":
                    price_changed = True

            if changed_fields:
                updated += 1
                db.add(
                    AuditLog(
                        user_id=user_id,
                        entity_type="car",
                        entity_id=car.id,
                        action=AuditAction.update,
                        old_value={
                            k: (str(v["old"]) if isinstance(v["old"], Decimal) else v["old"])
                            for k, v in changed_fields.items()
                        },
                        new_value={
                            k: (str(v["new"]) if isinstance(v["new"], Decimal) else v["new"])
                            for k, v in changed_fields.items()
                        },
                    )
                )
                if price_changed:
                    db.add(
                        AuditLog(
                            user_id=user_id,
                            entity_type="car",
                            entity_id=car.id,
                            action=AuditAction.price_change,
                            old_value={"price": str(old_price)},
                            new_value={"price": str(price)},
                        )
                    )
        else:
            car = Car(
                brand=incoming["brand"],
                model=incoming["model"],
                year=year,
                price=price,
                mileage=mileage,
                color=incoming.get("color"),
                location=incoming.get("location"),
                body_status=incoming.get("body_status"),
                description=incoming.get("description"),
                status=CarStatus.pending,
                source=CarSource.excel,
                excel_row_id=incoming.get("excel_row_id"),
            )
            db.add(car)
            await db.flush()
            added += 1
            db.add(
                AuditLog(
                    user_id=user_id,
                    entity_type="car",
                    entity_id=car.id,
                    action=AuditAction.create,
                    new_value={k: (str(v) if isinstance(v, Decimal) else v) for k, v in incoming.items()},
                )
            )

    # 2. Removed — archive, never hard delete
    for key, car in by_key.items():
        if key in rows_by_key:
            continue
        if car.status == CarStatus.archived:
            continue
        old_status = car.status
        car.status = CarStatus.archived
        removed += 1
        db.add(
            AuditLog(
                user_id=user_id,
                entity_type="car",
                entity_id=car.id,
                action=AuditAction.archive,
                old_value={"status": old_status.value},
                new_value={"status": CarStatus.archived.value, "reason": "excel_sync"},
            )
        )

    log_entry = ExcelImportLog(
        file_path=file_path,
        imported_rows=len(rows),
        added_rows=added,
        updated_rows=updated,
        removed_rows=removed,
        warnings=payload["diff"].get("warnings", []),
        applied_by_user_id=user_id,
        applied_at=datetime.now(tz=timezone.utc),
    )
    db.add(log_entry)
    await db.commit()
    await clear_diff(token)
    await invalidate_cars_cache()

    # Post-apply notifications
    from app.workers.tasks import notify_task

    if added:
        notify_task.delay(
            title="Excel sync — new cars added",
            body=f"{added} new cars created with status=pending. Please review.",
            meta={"added": added},
        )
    if removed:
        notify_task.delay(
            title="Excel sync — cars archived",
            body=f"{removed} cars were archived because they are no longer in the Excel.",
            meta={"removed": removed},
        )

    return {"added": added, "updated": updated, "removed": removed, "log_id": log_entry.id}
