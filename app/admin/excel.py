import io

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.admin._deps import admin_user_or_redirect
from app.database import get_db
from app.models import ExcelImportLog, UserRole
from app.services import excel_sync, media

router = APIRouter(prefix="/admin/excel", tags=["admin"], include_in_schema=False)


@router.get("")
async def excel_page(
    request: Request,
    user_or_redirect=Depends(admin_user_or_redirect),
    db: AsyncSession = Depends(get_db),
):
    if isinstance(user_or_redirect, RedirectResponse):
        return user_or_redirect
    logs = (
        await db.execute(select(ExcelImportLog).order_by(ExcelImportLog.created_at.desc()).limit(20))
    ).scalars().all()
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request,
        "admin/excel_import.html",
        {"user": user_or_redirect, "logs": logs, "error": None},
    )


@router.post("/upload")
async def excel_upload(
    request: Request,
    file: UploadFile,
    user_or_redirect=Depends(admin_user_or_redirect),
    db: AsyncSession = Depends(get_db),
):
    if isinstance(user_or_redirect, RedirectResponse):
        return user_or_redirect
    if user_or_redirect.role not in (UserRole.admin, UserRole.operator):
        raise HTTPException(status_code=403)
    path = await media.save_excel_upload(file)
    templates = request.app.state.templates
    try:
        diff, token = await excel_sync.parse_and_stage(db, path)
    except ValueError as exc:
        logs = (
            await db.execute(select(ExcelImportLog).order_by(ExcelImportLog.created_at.desc()).limit(20))
        ).scalars().all()
        return templates.TemplateResponse(
            request,
            "admin/excel_import.html",
            {"user": user_or_redirect, "logs": logs, "error": str(exc)},
            status_code=400,
        )
    return RedirectResponse(url=f"/admin/excel/preview/{token}", status_code=303)


@router.get("/preview/{token}")
async def excel_preview(
    request: Request,
    token: str,
    user_or_redirect=Depends(admin_user_or_redirect),
):
    if isinstance(user_or_redirect, RedirectResponse):
        return user_or_redirect
    payload = await excel_sync.load_diff(token)
    if not payload:
        raise HTTPException(status_code=404, detail="Preview expired")
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request,
        "admin/excel_preview.html",
        {
            "user": user_or_redirect,
            "token": token,
            "diff": payload["diff"],
            "file_path": payload["file_path"],
        },
    )


@router.get("/sample")
async def excel_sample(user_or_redirect=Depends(admin_user_or_redirect)):
    if isinstance(user_or_redirect, RedirectResponse):
        return user_or_redirect
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "خودروها"
    ws.sheet_view.rightToLeft = True

    headers = ["نام خودرو", "مدل", "کارکرد", "رنگ", "قیمت", "تامین کننده"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 20

    sample_rows = [
        ["پژو", "206 تیپ 5", 45000, "سفید", 850000000, "شرکت الف"],
        ["سمند", "LX", 80000, "نقره‌ای", 620000000, "نمایندگی ب"],
        ["تویوتا", "کرولا 2020", 12000, "مشکی", 1950000000, "واردات پارس"],
    ]
    row_font = Font(size=10)
    for row_data in sample_rows:
        row = ws.append(row_data)  # type: ignore[arg-type]

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = row_font
            cell.alignment = Alignment(horizontal="right")

    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers_resp = {
        "Content-Disposition": 'attachment; filename="sample_cars.xlsx"',
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


@router.post("/apply/{token}")
async def excel_apply(
    token: str,
    user_or_redirect=Depends(admin_user_or_redirect),
    db: AsyncSession = Depends(get_db),
):
    if isinstance(user_or_redirect, RedirectResponse):
        return user_or_redirect
    if user_or_redirect.role not in (UserRole.admin, UserRole.operator):
        raise HTTPException(status_code=403)
    try:
        await excel_sync.apply_diff(db, token, user_or_redirect.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return RedirectResponse(url="/admin/excel", status_code=303)
